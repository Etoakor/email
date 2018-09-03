# encoding=utf-8
import traceback

import pymysql as pms
import openpyxl
import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib

def get_datas(sql):
    # 一个传入sql导出数据的函数
    # 跟数据库建立连接
    conn = pms.connect(host='localhost', user='用户',
                       passwd='数据库密码', database='csv', port=3306, charset="utf8")
    # 使用 cursor() 方法创建一个游标对象 cursor
    cur = conn.cursor()
    # 使用 execute() 方法执行 SQL
    cur.execute(sql)
    # 获取所需要的数据
    datas = cur.fetchall()
    #关闭连接
    cur.close()
    #返回所需的数据
    return datas

def get_fields(sql):
    # 一个传入sql导出字段的函数
    conn = pms.connect(host='localhost', user='用户',
                       passwd='数据库密码', database='csv', port=3306, charset="utf8")
    cur = conn.cursor()
    cur.execute(sql)
    # 获取所需要的字段名称
    fields = cur.description
    cur.close()
    return fields

def get_excel(data, field, file):
    # 将数据和字段名写入excel的函数
    #新建一个工作薄对象
    new = openpyxl.Workbook()
    #激活一个新的sheet
    sheet = new.active
    #给sheet命名
    sheet.title = '数据'
    #将字段名称循环写入excel第一行，因为字段格式列表里包含列表，每个列表的第一元素才是字段名称
    for col in range(len(field)):
        #row代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始，这点于python不同要注意
        _ = sheet.cell(row=1, column=col+1, value=u'%s' % field[col][0])
     #将数据循环写入excel的每个单元格中
    for row in range(len(data)):
        for col in range(len(field)):
            #因为第一行写了字段名称，所以要从第二行开始写入
            _ = sheet.cell(row=row+2, column=col + 1, value=u'%s' % data[row][col])
            #将生成的excel保存，这步是必不可少的
    newworkbook = new.save(file)
    #返回生成的excel
    return newworkbook

def getYesterday():
    # 获取昨天日期的字符串格式的函数
    #获取今天的日期
    today = datetime.date.today()
    #获取一天的日期格式数据
    oneday = datetime.timedelta(days=1)
    #昨天等于今天减去一天
    yesterday = today - oneday
    #获取昨天日期的格式化字符串
    yesterdaystr = yesterday.strftime('%Y-%m-%d')
    #返回昨天的字符串
    return yesterdaystr

def create_email(email_from, email_to, email_Subject, email_text, annex_path, annex_name):
    # 输入发件人昵称、收件人昵称、主题，正文，附件地址,附件名称生成一封邮件
    #生成一个空的带附件的邮件实例
    message = MIMEMultipart()
    #将正文以text的形式插入邮件中
    message.attach(MIMEText("易途人", 'plain', 'utf-8'))
    #生成发件人名称（这个跟发送的邮件没有关系）
    message['From'] = Header("易途", 'utf-8')
    #生成收件人名称（这个跟接收的邮件也没有关系）
    message['To'] = Header(email_to, 'utf-8')
    #生成邮件主题
    message['Subject'] = Header(email_Subject, 'utf-8')
    #读取附件的内容
    att1 = MIMEText(open(annex_path, 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    #生成附件的名称
    att1["Content-Disposition"] = 'attachment; filename=' + annex_name
    #将附件内容插入邮件中
    message.attach(att1)
    #返回邮件
    return message

def send_email(sender, password, receiver, msg):
    # 一个输入邮箱、密码、收件人、邮件内容发送邮件的函数
    try:
        #找到你的发送邮箱的服务器地址，已加密的形式发送
        server = smtplib.SMTP_SSL("smtp.aliyun.com", 465)  # 发件人邮箱中的SMTP服务器
        server.ehlo()
        #登录你的账号
        server.login(sender, password)  # 括号中对应的是发件人邮箱账号、邮箱密码
        #发送邮件
        server.sendmail(sender, receiver, msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号（是一个列表）、邮件内容
        print("邮件发送成功")
        server.quit()  # 关闭连接
    except Exception:
        print(traceback.print_exc())
        print("邮件发送失败")

def main():
    print(datetime.datetime.now())
    my_sql = sql = "SELECT id '用户ID',\
                          number '手机号',\
                          area '地区'\
                          FROM csv;"
    # 生成数据
    my_data = get_datas(my_sql)
    # 生成字段名称
    my_field = get_fields(my_sql)
    # 得到昨天的日期
    yesterdaystr = getYesterday()
    # 文件名称
    my_file_name = 'etoakor'+yesterdaystr + '.xlsx'
    # 文件路径
    file_path = '/Users/zhangxianan/Desktop/1/' + my_file_name
    # 生成excel
    get_excel(my_data, my_field, file_path)

    my_email_from = '易途人'
    my_email_to = '易途'
    # 邮件标题
    my_email_Subject = 'Etoakor' + yesterdaystr
    # 邮件正文
    my_email_text = "易途：\n\t用户数据，请查收！\n\n 易途人团队"
    #附件地址
    my_annex_path = file_path
    #附件名称
    my_annex_name = my_file_name
    # 生成邮件
    my_msg = create_email(my_email_from, my_email_to, my_email_Subject,
                          my_email_text, my_annex_path, my_annex_name)
    my_sender = 'etoakor@aliyun.com'
    my_password = '密码'
    my_receiver = ['xuan2011long@sina.com']#接收人邮箱列表
    # 发送邮件
    send_email(my_sender, my_password, my_receiver, my_msg)
    print(datetime.datetime.now())

if __name__ == "__main__":
    main();